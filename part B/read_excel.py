
from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet.title = "lang"
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
code = ['KA', 'TS', 'TN']

sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i-2]
    sheet.cell(row=i, column=2).value = lang[i-2]
    sheet.cell(row=i, column=3).value = code[i-2]
wb.save("C:\\Users\\sindh\\OneDrive\\Desktop\\VS PYTHON\\demo.xlsx")

for row in range(0, sheet.max_row):
	for col in sheet.iter_cols(1, sheet.max_column):
		print(col[row].value,end=" ")
	print()
